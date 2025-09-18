
import io
from io import BytesIO
import pandas as pd
import numpy as np
import streamlit as st
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Masterfile Filler", layout="wide")

st.markdown("<h1 style='text-align: center;'>🧩 Masterfile Filler</h1>", unsafe_allow_html=True)

st.markdown("<h4 style='text-align: center; font-style: italic;'>Innovating with AI Today ⏐ Leading Automation Tomorrow</h4>", unsafe_allow_html=True)
# ---------- Helpers ----------

def read_tabular(uploaded_file, sheet: Optional[str] = None) -> pd.DataFrame:
    """Read Excel/CSV into DataFrame. If Excel and sheet is None, use first sheet."""
    if uploaded_file is None:
        return pd.DataFrame()
    name = uploaded_file.name.lower()
    if name.endswith(('.csv', '.txt')):
        return pd.read_csv(uploaded_file)
    else:
        # Excel: read specified sheet or the first
        try:
            if sheet is not None:
                return pd.read_excel(uploaded_file, sheet_name=sheet, engine="openpyxl")
            # default first sheet
            xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
            first = xls.sheet_names[0]
            return pd.read_excel(xls, sheet_name=first, engine="openpyxl")
        except Exception as e:
            st.error(f"Failed to read table: {e}")
            return pd.DataFrame()


def read_mapping(uploaded_file) -> pd.DataFrame:
    """Read mapping CSV/Excel and normalize columns."""
    if uploaded_file is None:
        return pd.DataFrame()
    try:
        name = uploaded_file.name.lower()
        if name.endswith(('.csv', '.txt')):
            m = pd.read_csv(uploaded_file)
        else:
            m = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        st.error(f"Failed to read mapping: {e}")
        return pd.DataFrame()
    # Normalize column names
    m.columns = [c.strip().lower() for c in m.columns]
    # Expected columns
    required_any = {'template_header', 'template_cell'}
    expected_base = {'template_sheet', 'raw_column', 'default_value', 'dtype', 'required', 'notes'}
    # Fill missing optional columns
    for col in list(expected_base | required_any):
        if col not in m.columns:
            m[col] = np.nan
    # Validate presence of template_sheet
    if 'template_sheet' not in m.columns:
        st.error("Mapping must include a 'template_sheet' column.")
        return pd.DataFrame()
    # Drop rows with neither template_header nor template_cell
    m = m[~(m['template_header'].isna() & m['template_cell'].isna())].copy()
    # Clean dtype entries
    m['dtype'] = m['dtype'].astype(str).str.lower().replace({'nan': ''})
    # Boolean 'required'
    def to_bool(x):
        s = str(x).strip().lower()
        return s in ('true','1','yes','y','required')
    m['required'] = m['required'].apply(to_bool)
    # Strip strings
    for c in ['template_sheet', 'template_header', 'template_cell', 'raw_column', 'default_value']:
        m[c] = m[c].astype(str).str.strip().replace({'nan': ''})
    return m


def coerce_dtype(val, dtype_str: str):
    if dtype_str in (None, '', 'str', 'text'):
        if pd.isna(val):
            return ''
        return str(val)
    try:
        if dtype_str in ('int', 'integer'):
            if pd.isna(val) or val == '':
                return None
            return int(float(val))
        if dtype_str in ('float', 'number', 'decimal'):
            if pd.isna(val) or val == '':
                return None
            return float(val)
        if dtype_str in ('bool', 'boolean'):
            if isinstance(val, (int, float)):
                return bool(val)
            s = str(val).strip().lower()
            return s in ('true','1','yes','y','t')
        if dtype_str in ('date','datetime'):
            if pd.isna(val) or val == '':
                return None
            return pd.to_datetime(val, errors='coerce')
        # default
        if pd.isna(val):
            return ''
        return str(val)
    except Exception:
        # Fallback to raw
        return val


def detect_header_row(ws, needed_headers: List[str], search_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    """Find the row index (1-based) that contains the template headers. Return row and mapping of header->col index."""
    nh = [h.strip().lower() for h in needed_headers if h and str(h).strip()]
    best_row = None
    best_match_count = -1
    header_map = {}
    max_row = min(ws.max_row or 1, search_rows)
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        val_to_idx = {}
        for idx, v in enumerate(row_vals, start=1):
            if v is None:
                continue
            key = str(v).strip().lower()
            if key and key not in val_to_idx:
                val_to_idx[key] = idx
        match_count = sum(1 for h in nh if h in val_to_idx)
        if match_count > best_match_count:
            best_match_count = match_count
            best_row = r
            header_map = {h: val_to_idx[h] for h in nh if h in val_to_idx}
        if match_count == len(nh) and len(nh) > 0:
            break
    return best_row or 1, header_map


def write_table_into_template(
    template_bytes: BytesIO,
    template_sheet: str,
    mapping_for_sheet: pd.DataFrame,
    raw_df: pd.DataFrame,
    auto_detect_headers: bool = True
) -> BytesIO:
    """Open the template workbook and write values according to mapping; return output bytes."""
    wb = load_workbook(filename=template_bytes)
    if template_sheet not in wb.sheetnames:
        ws = wb.create_sheet(template_sheet)
        header_row_idx = 1
        header_map = {}
    else:
        ws = wb[template_sheet]
        needed_headers = [x for x in mapping_for_sheet['template_header'].tolist() if x]
        if auto_detect_headers and len(needed_headers) > 0:
            header_row_idx, header_map = detect_header_row(ws, needed_headers)
        else:
            header_row_idx = 1
            header_map = {}

    mapping_headers = [h for h in mapping_for_sheet['template_header'].tolist() if h]
    if len(mapping_headers) > 0:
        current_headers = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        current_headers_norm = [str(x).strip().lower() if x is not None else '' for x in current_headers]
        used_cols = set(header_map.values())
        next_free_col = 1
        while next_free_col in used_cols:
            next_free_col += 1
        for h in mapping_headers:
            key = h.strip().lower()
            if key in current_headers_norm:
                col_idx = current_headers_norm.index(key) + 1
                header_map[key] = col_idx
            elif key in header_map:
                pass
            else:
                while next_free_col in used_cols:
                    next_free_col += 1
                ws.cell(row=header_row_idx, column=next_free_col, value=h)
                header_map[key] = next_free_col
                used_cols.add(next_free_col)
                next_free_col += 1

        start_row = header_row_idx + 1
        maps = []
        for _, row in mapping_for_sheet.iterrows():
            th = str(row.get('template_header', '')).strip()
            if not th:
                continue
            key = th.lower()
            col_idx = header_map.get(key)
            if not col_idx:
                continue
            maps.append({
                'template_header': th,
                'col_idx': col_idx,
                'raw_column': str(row.get('raw_column','')).strip(),
                'default_value': row.get('default_value',''),
                'dtype': str(row.get('dtype','')).strip().lower(),
                'required': bool(row.get('required', False))
            })

        errs = []
        for m in maps:
            if m['required'] and m['raw_column'] and m['raw_column'] not in raw_df.columns:
                errs.append(f"Required raw column '{m['raw_column']}' not found for template header '{m['template_header']}'.")
        if errs:
            st.warning("\n".join(errs))

        for i, (_, r) in enumerate(raw_df.iterrows()):
            for m in maps:
                if m['raw_column'] and m['raw_column'] in raw_df.columns:
                    val = r[m['raw_column']]
                    if pd.isna(val) or val == '':
                        val = m['default_value']
                else:
                    val = m['default_value']
                out_val = coerce_dtype(val, m['dtype'])
                ws.cell(row=start_row + i, column=m['col_idx'], value=out_val)

        try:
            for key, col_idx in header_map.items():
                max_width = len(str(ws.cell(row=header_row_idx, column=col_idx).value or ''))
                for rr in range(start_row, start_row + min(len(raw_df), 200)):
                    v = ws.cell(row=rr, column=col_idx).value
                    max_width = max(max_width, len(str(v)) if v is not None else 0)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 60)
        except Exception:
            pass

    cell_rows = mapping_for_sheet[~mapping_for_sheet['template_cell'].isna() & (mapping_for_sheet['template_cell']!='')]
    if not cell_rows.empty:
        for _, row in cell_rows.iterrows():
            cell_addr = str(row['template_cell']).strip()
            raw_col = str(row.get('raw_column','')).strip()
            default_val = row.get('default_value','')
            dtype = str(row.get('dtype','')).strip().lower()
            if default_val not in (None, '', 'nan'):
                val = default_val
            elif raw_col and raw_col in raw_df.columns and len(raw_df) > 0:
                series = raw_df[raw_col].dropna()
                val = series.iloc[0] if not series.empty else ''
            else:
                val = ''
            out_val = coerce_dtype(val, dtype)
            try:
                ws[cell_addr] = out_val
            except Exception as e:
                st.warning(f"Could not write to cell {cell_addr}: {e}")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ---------- SESSION ----------

if 'raw_df' not in st.session_state:
    st.session_state.raw_df = pd.DataFrame()
if 'raw_sheet' not in st.session_state:
    st.session_state.raw_sheet = None
if 'template_bytes' not in st.session_state:
    st.session_state.template_bytes = None
if 'template_sheet' not in st.session_state:
    st.session_state.template_sheet = None
if 'mapping_df' not in st.session_state:
    st.session_state.mapping_df = pd.DataFrame()

tab1, tab2, tab3 = st.tabs(["1) Upload Raw Data", "2) Upload Template (Masterfile)", "3) Process & Download"])

# ---- Tab 1: Raw Data ----
with tab1:
    st.subheader("Upload Raw Data (CSV or Excel)")
    raw_file = st.file_uploader("Choose raw data file", type=["csv","xlsx","xls"], key="raw_uploader")
    raw_sheet = None
    if raw_file is not None and raw_file.name.lower().endswith(('xlsx','xls')):
        try:
            xls = pd.ExcelFile(raw_file, engine="openpyxl")
            raw_sheet = st.selectbox("Select sheet", xls.sheet_names, index=0)
            st.session_state.raw_sheet = raw_sheet
            df_prev = pd.read_excel(xls, sheet_name=raw_sheet, engine="openpyxl")
        except Exception as e:
            st.error(f"Could not read Excel: {e}")
            df_prev = pd.DataFrame()
    else:
        df_prev = read_tabular(raw_file)

    if raw_file is not None:
        st.session_state.raw_df = df_prev
        st.success(f"Loaded {len(df_prev):,} rows × {len(df_prev.columns):,} columns.")
        st.dataframe(df_prev.head(50), use_container_width=True)

# ---- Tab 2: Template ----
with tab2:
    st.subheader("Upload Marketplace Template (Excel)")
    template_file = st.file_uploader("Choose template (masterfile) Excel", type=["xlsx","xls"], key="template_uploader")
    if template_file is not None:
        template_bytes = BytesIO(template_file.getbuffer())
        st.session_state.template_bytes = template_bytes
        try:
            xls = pd.ExcelFile(template_file, engine="openpyxl")
            sheet_name = st.selectbox("Select template sheet to fill", xls.sheet_names, index=0)
        except Exception as e:
            st.error(f"Failed to open template: {e}")
            sheet_name = None
        st.session_state.template_sheet = sheet_name
        if sheet_name:
            prev = pd.read_excel(template_file, sheet_name=sheet_name, engine="openpyxl", nrows=5)
            st.caption("Preview (first 5 rows):")
            st.dataframe(prev, use_container_width=True)

# ---- Tab 3: Process & Download ----
with tab3:
    st.subheader("Mapping & Processing")
    st.markdown(
        "**Mapping file format (Excel/CSV)** — include these columns:\n"
        "- `template_sheet` *(required)*: Sheet name in the template workbook (e.g., `Masterfile`)  \n"
        "- `template_header` *(row-wise)*: Column header text where data should go (match header text in the template)  \n"
        "- `template_cell` *(one-off)*: Excel cell address for single values (e.g., `B2`)  \n"
        "- `raw_column`: Column from your raw data to copy from  \n"
        "- `default_value`: Fallback if raw data is missing/blank  \n"
        "- `dtype`: Optional conversion (`str`, `int`, `float`, `bool`, `date`)  \n"
        "- `required`: Mark `true` if the `raw_column` must exist in raw data"
    )
    mapping_file = st.file_uploader("Upload mapping file (xlsx/csv)", type=["xlsx","xls","csv"], key="mapping_uploader")
    autodetect = st.checkbox("Auto-detect header row in template", value=True)
    if mapping_file is not None:
        mapping_df = read_mapping(mapping_file)
        st.session_state.mapping_df = mapping_df
        if not mapping_df.empty:
            st.success(f"Loaded mapping with {len(mapping_df):,} lines.")
            st.dataframe(mapping_df.head(100), use_container_width=True)
        else:
            st.warning("Mapping is empty or invalid.")

    st.divider()
    can_process = (
        not st.session_state.raw_df.empty
        and st.session_state.template_bytes is not None
        and st.session_state.template_sheet is not None
        and not st.session_state.mapping_df.empty
    )
    if not can_process:
        st.info("Please upload raw data, template, and mapping to enable processing.")

    if st.button("⚙️ Process & Download", type="primary", disabled=not can_process):
        try:
            m = st.session_state.mapping_df
            template_bytes = BytesIO(st.session_state.template_bytes.getbuffer())
            sheets_in_map = [s for s in m['template_sheet'].dropna().astype(str).unique() if s]
            if len(sheets_in_map) == 0 and st.session_state.template_sheet:
                sheets_in_map = [st.session_state.template_sheet]
                m['template_sheet'] = st.session_state.template_sheet

            out_bytes = None
            for s in sheets_in_map:
                ms = m[m['template_sheet'].astype(str) == s].copy()
                out_bytes = write_table_into_template(
                    template_bytes = template_bytes if out_bytes is None else out_bytes,
                    template_sheet = s,
                    mapping_for_sheet = ms,
                    raw_df = st.session_state.raw_df,
                    auto_detect_headers = autodetect
                )
            if out_bytes is None:
                st.error("Nothing was processed. Check your mapping.")
            else:
                st.success("Done! Your filled masterfile is ready to download.")
                st.download_button(
                    label="⬇️ Download Filled Masterfile (Excel)",
                    data=out_bytes.getvalue(),
                    file_name="filled_masterfile.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.exception(e)
